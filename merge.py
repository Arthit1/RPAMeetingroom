from pathlib import Path
import pandas as pd
import re
import sys
import math
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

DOWNLOAD_DIR = Path("downloads")
OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "merged_report.xlsx"
DEBUG_RAW_FILE = OUTPUT_DIR / "debug_raw_concat.xlsx"
DEBUG_SUMMARY_FILE = OUTPUT_DIR / "debug_summary.xlsx"

OUTPUT_DIR.mkdir(exist_ok=True)

# =========================
# รับวันที่จาก command line
# =========================
if len(sys.argv) >= 7:
    START_DAY = int(sys.argv[1])
    START_MONTH = int(sys.argv[2])
    START_YEAR = int(sys.argv[3])
    END_DAY = int(sys.argv[4])
    END_MONTH = int(sys.argv[5])
    END_YEAR = int(sys.argv[6])
else:
    START_DAY = 1
    START_MONTH = 1
    START_YEAR = 2026
    END_DAY = 31
    END_MONTH = 1
    END_YEAR = 2026

# =========================
# ตัวแปรปรับได้
# =========================
HOURS_PER_DAY = 8.5
NORMAL_OPEN_DAYS = 19
FLOOR22_OPEN_DAYS = 22
B_FLOOR_OPEN_DAYS_DEFAULT = 26

B_FLOOR_OPEN_DAYS_MAP = {
    "B101": 20,
    "B102": 20,
    "B103": 20,
    "B104": 21,
    "B105": 22,
    "B106": 20,
    "B107": 20,
    "B108": 21,
}

FINAL_COLUMNS = [
    "No",
    "Room",
    "Number of Bookings",
    "Number of Meet Now",
    "Number of Advanced Bookings",
    "Total Duration",
    "Total Duration (Hours)",
    "Avg Duration",
    "Min Duration",
    "Max Duration",
    "Open Days",
    "Available Hours",
    "Utilization (%)",
    "Source File",
    "Source Sheet"
]

EXPECTED_OUTPUT_COLUMNS = [
    "No",
    "Room",
    "Number of Bookings",
    "Number of Meet Now",
    "Number of Advanced Bookings",
    "Total Duration",
    "Avg Duration",
    "Min Duration",
    "Max Duration"
]

def clean_room(room):
    if pd.isna(room):
        return ""
    text = str(room).strip().upper().replace(".", "")
    text = text.replace(" ", "")
    text = text.replace("-", "")
    return text

def parse_room_key(room):
    if pd.isna(room):
        return (9, 999999, 999999, "")

    text = str(room).strip().replace(".", "")
    upper_text = text.upper()

    m = re.match(r"^(\d+)-(\d+)$", upper_text)
    if m:
        return (0, int(m.group(1)), int(m.group(2)), upper_text)

    m = re.match(r"^([A-Z]+)(\d+)-(\d+)$", upper_text)
    if m:
        return (1, int(m.group(2)), int(m.group(3)), m.group(1))

    nums = re.findall(r"\d+", upper_text)
    if nums:
        first = int(nums[0])
        second = int(nums[1]) if len(nums) > 1 else 0
        return (2, first, second, upper_text)

    return (3, 999999, 999999, upper_text)

def duration_to_hours(value):
    if pd.isna(value):
        return 0

    text = str(value).strip()
    if text == "":
        return 0

    m = re.search(r"(?:(\d+)\s+days?\s+)?(\d{1,2}):(\d{2})(?::(\d{2}))?", text)
    if m:
        days = int(m.group(1)) if m.group(1) else 0
        hours = int(m.group(2))
        minutes = int(m.group(3))
        seconds = int(m.group(4)) if m.group(4) else 0
        total_hours = (days * 24) + hours + (minutes / 60) + (seconds / 3600)
        return math.ceil(total_hours)

    try:
        return math.ceil(float(text))
    except Exception:
        return 0

def get_open_days(room):
    room_key = clean_room(room)

    if room_key.startswith("B"):
        return B_FLOOR_OPEN_DAYS_MAP.get(room_key, B_FLOOR_OPEN_DAYS_DEFAULT)

    if room_key.startswith("22"):
        return FLOOR22_OPEN_DAYS

    return NORMAL_OPEN_DAYS

def looks_like_header_row(row_values):
    row_text = [str(v).strip().lower() for v in row_values]
    joined = " | ".join(row_text)
    keywords = [
        "room",
        "number of bookings",
        "number of meet now",
        "number of advanced bookings",
        "total duration",
        "avg duration",
        "min duration",
        "max duration",
    ]
    hits = sum(1 for k in keywords if k in joined)
    return hits >= 3

def normalize_sheet(df, source_file, source_sheet):
    """
    พยายาม normalize โดยไม่ทิ้งข้อมูลเกินจำเป็น
    """
    original_rows = len(df)

    # ลบแถวว่างทั้งแถว
    df = df.dropna(how="all").copy()
    after_dropna = len(df)

    # reset index
    df = df.reset_index(drop=True)

    # ถ้าไฟล์มีหัว 5 แถวตามแพทเทิร์นเดิม ให้ตัด
    # แต่ถ้าไม่ใช่ จะไม่ตัดมั่ว
    start_idx = 0
    for i in range(min(10, len(df))):
        row_vals = df.iloc[i].tolist()
        if looks_like_header_row(row_vals):
            start_idx = i + 1
            break

    if start_idx > 0:
        df = df.iloc[start_idx:].copy().reset_index(drop=True)

    # จำกัดคอลัมน์อย่างปลอดภัย
    if df.shape[1] < 9:
        return None, {
            "source_file": source_file,
            "source_sheet": source_sheet,
            "original_rows": original_rows,
            "after_dropna": after_dropna,
            "final_rows": 0,
            "status": f"skip: columns<{9}"
        }

    df = df.iloc[:, :9].copy()
    df.columns = EXPECTED_OUTPUT_COLUMNS

    # ลบแถว header ซ้ำที่ติดมากลางไฟล์
    header_mask = df.apply(lambda r: looks_like_header_row(r.tolist()), axis=1)
    df = df[~header_mask].copy()

    # ลบเฉพาะแถวที่ Room ว่างจริง
    df["Room"] = df["Room"].astype(str).str.strip()
    df = df[df["Room"].notna()].copy()
    df = df[df["Room"].astype(str).str.strip() != ""].copy()
    df = df[df["Room"].astype(str).str.lower() != "nan"].copy()

    df["Source File"] = source_file
    df["Source Sheet"] = source_sheet

    info = {
        "source_file": source_file,
        "source_sheet": source_sheet,
        "original_rows": original_rows,
        "after_dropna": after_dropna,
        "final_rows": len(df),
        "status": "ok"
    }
    return df, info

files = sorted(DOWNLOAD_DIR.glob("*.xlsx"))

if not files:
    print("ไม่พบไฟล์ Excel")
    raise SystemExit

all_data = []
debug_summary = []

for file in files:
    try:
        xl = pd.ExcelFile(file)
        for sheet_name in xl.sheet_names:
            raw_df = pd.read_excel(file, sheet_name=sheet_name, header=None)
            normalized_df, info = normalize_sheet(raw_df, file.name, sheet_name)
            debug_summary.append(info)

            if normalized_df is not None and not normalized_df.empty:
                all_data.append(normalized_df)

            print(
                f"อ่านไฟล์ {file.name} | sheet={sheet_name} | "
                f"rows_before={info['original_rows']} | rows_after={info['final_rows']} | {info['status']}"
            )
    except Exception as e:
        debug_summary.append({
            "source_file": file.name,
            "source_sheet": "",
            "original_rows": 0,
            "after_dropna": 0,
            "final_rows": 0,
            "status": f"error: {e}"
        })
        print(f"อ่านไฟล์ไม่สำเร็จ {file.name}: {e}")

if not all_data:
    print("ไม่มีข้อมูลที่รวมได้")
    pd.DataFrame(debug_summary).to_excel(DEBUG_SUMMARY_FILE, index=False)
    raise SystemExit

merged = pd.concat(all_data, ignore_index=True)

# เก็บไฟล์ raw concat สำหรับ debug
merged.to_excel(DEBUG_RAW_FILE, index=False)

# sort
merged["_sort"] = merged["Room"].apply(parse_room_key)
merged = merged.sort_values("_sort", kind="stable").drop(columns="_sort")

# reset No
merged = merged.drop(columns=["No"])
merged.insert(0, "No", range(1, len(merged) + 1))

# คำนวณ
merged["Total Duration (Hours)"] = merged["Total Duration"].apply(duration_to_hours)
merged["Open Days"] = merged["Room"].apply(get_open_days)
merged["Available Hours"] = merged["Open Days"] * HOURS_PER_DAY
merged["Utilization (%)"] = (
    merged["Total Duration (Hours)"] / merged["Available Hours"] * 100
).round(0).fillna(0).astype(int).astype(str) + "%"

merged = merged[FINAL_COLUMNS]

report_title = (
    f"ข้อมูลการใช้งานห้องประชุม วันที่ "
    f"{START_DAY:02d}/{START_MONTH:02d}/{START_YEAR} "
    f"ถึง {END_DAY:02d}/{END_MONTH:02d}/{END_YEAR}"
)

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    merged.to_excel(writer, index=False, startrow=2)
    pd.DataFrame(debug_summary).to_excel(writer, sheet_name="Debug Summary", index=False)

    worksheet = writer.sheets["Sheet1"]
    total_columns = len(merged.columns)
    last_column = get_column_letter(total_columns)

    worksheet.merge_cells(f"A1:{last_column}1")
    worksheet["A1"] = report_title
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet.row_dimensions[1].height = 24

    for cell in worksheet[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

print("รวมไฟล์เสร็จแล้ว")
print("ไฟล์อยู่ที่:", OUTPUT_FILE)
print("ไฟล์ debug raw อยู่ที่:", DEBUG_RAW_FILE)
print("ไฟล์ debug summary อยู่ที่:", DEBUG_SUMMARY_FILE)